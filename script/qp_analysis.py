import argparse
import os
import re
import subprocess
import time
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter


@dataclass
class EncodeResult:
    qp: int
    frames: int
    fps: float
    bitrate_kbps: float
    out_file: str
    out_size_bytes: int
    elapsed_sec: float
    return_code: int


# Example line:
# "encoded 300 frames, 182.82 fps, 511.45 kb/s"
ENCODE_SUMMARY_RE = re.compile(
    r"encoded\s+(?P<frames>\d+)\s+frames,\s+"
    r"(?P<fps>\d+(\.\d+)?)\s+fps,\s+"
    r"(?P<kbps>\d+(\.\d+)?)\s+kb/s",
    re.IGNORECASE
)


def run_x264_once(x264_path: Path, input_path: Path, input_res: str, qp: int, out_dir: Path) -> EncodeResult:
    out_file = out_dir / f"foreman_qp{qp:02d}.264"

    cmd = [
        str(x264_path),
        "--qp", str(qp),
        "--input-res", input_res,
        "-o", str(out_file),
        str(input_path)
    ]

    start = time.perf_counter()
    proc = subprocess.run(cmd, capture_output=True, text=True)
    elapsed = time.perf_counter() - start

    # x264 tends to print progress to stderr, summary may appear in stderr or stdout depending on build
    combined = (proc.stdout or "") + "\n" + (proc.stderr or "")

    m = ENCODE_SUMMARY_RE.search(combined)
    if not m:
        raise RuntimeError(
            f"Could not parse x264 summary line for QP={qp}.\n"
            f"Command: {' '.join(cmd)}\n"
            f"Output (truncated):\n{combined[-2000:]}"
        )

    frames = int(m.group("frames"))
    fps = float(m.group("fps"))
    kbps = float(m.group("kbps"))

    size_bytes = out_file.stat().st_size if out_file.exists() else 0

    return EncodeResult(
        qp=qp,
        frames=frames,
        fps=fps,
        bitrate_kbps=kbps,
        out_file=str(out_file),
        out_size_bytes=size_bytes,
        elapsed_sec=elapsed,
        return_code=proc.returncode
    )


def write_csv(results: List[EncodeResult], csv_path: Path) -> None:
    import csv
    fieldnames = list(asdict(results[0]).keys()) if results else []
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in results:
            w.writerow(asdict(r))


def autosize_columns(ws) -> None:
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def add_line_chart(ws, title: str, x_col: int, y_col: int, start_row: int, end_row: int, anchor: str) -> None:
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = ws.cell(row=1, column=y_col).value
    chart.x_axis.title = ws.cell(row=1, column=x_col).value

    xvalues = Reference(ws, min_col=x_col, min_row=start_row, max_row=end_row)
    yvalues = Reference(ws, min_col=y_col, min_row=1, max_row=end_row)  # include header
    chart.add_data(yvalues, titles_from_data=True)
    chart.set_categories(xvalues)
    chart.height = 11
    chart.width = 22
    ws.add_chart(chart, anchor)


def write_excel_report(results: List[EncodeResult], xlsx_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QP Results"

    headers = [
        "qp", "frames", "fps", "bitrate_kbps",
        "out_size_bytes", "out_size_kb", "out_size_mb",
        "elapsed_sec", "return_code", "out_file"
    ]
    ws.append(headers)

    for r in results:
        kb = r.out_size_bytes / 1024.0
        mb = r.out_size_bytes / (1024.0 * 1024.0)
        ws.append([
            r.qp, r.frames, r.fps, r.bitrate_kbps,
            r.out_size_bytes, round(kb, 3), round(mb, 6),
            round(r.elapsed_sec, 4), r.return_code, r.out_file
        ])

    autosize_columns(ws)

    start_row = 2
    end_row = 1 + len(results)

    # Charts: QP vs FPS, QP vs Output Size (KB), QP vs Bitrate
    add_line_chart(ws, "QP vs Encoding Speed (FPS)", x_col=1, y_col=3,
                   start_row=start_row, end_row=end_row, anchor="L2")
    add_line_chart(ws, "QP vs Output Size (KB)", x_col=1, y_col=6,
                   start_row=start_row, end_row=end_row, anchor="L20")
    add_line_chart(ws, "QP vs Bitrate (kb/s)", x_col=1, y_col=4,
                   start_row=start_row, end_row=end_row, anchor="L38")

    wb.save(xlsx_path)


def main():
    p = argparse.ArgumentParser(description="Analyze x264 QP 1-51: FPS and output size.")
    p.add_argument("--x264", required=True, help="Path to x264 executable (e.g., ./x264 or x264.exe)")
    p.add_argument("--input", required=True, help="Path to input .yuv file (e.g., foreman-cif.yuv)")
    p.add_argument("--res", required=True, help="Input resolution like 352x288")
    p.add_argument("--qp-min", type=int, default=11)
    p.add_argument("--qp-max", type=int, default=51)
    p.add_argument("--out-dir", default="out_qp", help="Output directory for encoded .264 files and reports")
    args = p.parse_args()

    x264_path = Path(args.x264).expanduser().resolve()
    input_path = Path(args.input).expanduser().resolve()
    out_dir = Path(args.out_dir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    if not x264_path.exists():
        raise FileNotFoundError(f"x264 not found at: {x264_path}")
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found at: {input_path}")

    results: List[EncodeResult] = []
    for qp in range(args.qp_min, args.qp_max + 1):
        print(f"Running QP={qp} ...")
        r = run_x264_once(x264_path, input_path, args.res, qp, out_dir)
        results.append(r)
        print(f"  -> frames={r.frames}, fps={r.fps:.2f}, kb/s={r.bitrate_kbps:.2f}, size={r.out_size_bytes} bytes")

    csv_path = out_dir / "qp_results.csv"
    xlsx_path = out_dir / "qp_report.xlsx"
    write_csv(results, csv_path)
    write_excel_report(results, xlsx_path)

    print("\nDone.")
    print(f"CSV:  {csv_path}")
    print(f"XLSX: {xlsx_path}")


if __name__ == "__main__":
    main()